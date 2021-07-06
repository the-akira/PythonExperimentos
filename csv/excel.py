from openpyxl.chart import BarChart, Reference
import openpyxl as xl

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # Imprime os headers do arquivo
    for i in range(2,9):
        headers = sheet.cell(1,i)
        print(headers.value)

    # Imprime todos os nomes da coluna First Name
    nomes = [sheet['b' + str(i)] for i in range(1,52)]
    for nome in nomes:
        print(nome.value)

    # Altera a idade (Coluna 6) e cria uma nova coluna de idades (Coluna 9)
    for row in range(2, sheet.max_row + 1):
        idades = sheet.cell(row, 6)
        idades_alteradas = idades.value - 8 
        idades_alteradas_cell = sheet.cell(row, 9)
        idades_alteradas_cell.value = idades_alteradas

    # Cria um gráfico de barras com a nova coluna de idades (Coluna 9)
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=9, max_col=9)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)

process_workbook('pessoas.xlsx')