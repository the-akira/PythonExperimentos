from flask import (
    Flask, 
    render_template, 
    request, 
    redirect, 
    Response, 
    jsonify, 
    json, 
    flash, 
    url_for
)
from openpyxl.utils import get_column_letter
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from io import BytesIO
import pandas as pd
import yaml

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///students.db'
app.config['JSON_AS_ASCII'] = False
db = SQLAlchemy(app)

class Classroom(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)
    time = db.Column(db.String(255), nullable=False)
    students = db.relationship('Student', backref='classroom', lazy='joined', cascade='all, delete-orphan')

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)
    classroom_id = db.Column(db.Integer, db.ForeignKey('classroom.id'), nullable=False)
    grades = db.relationship('Grade', backref='student', lazy='joined', cascade='all, delete-orphan')

class Grade(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    value = db.Column(db.Integer, nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)

with app.app_context():
    db.create_all()

@app.route('/')
def index():
    classrooms = Classroom.query.all()
    return render_template('index.html', classrooms=classrooms)

@app.route('/add_classroom', methods=['GET', 'POST'])
def add_classroom():
    if request.method == 'POST':
        name = request.form['name']
        time = request.form['time']
        classroom = Classroom(name=name, time=time)
        db.session.add(classroom)
        db.session.commit()
        return redirect('/')
    return render_template('add_classroom.html')

@app.route('/delete_classroom/<int:classroom_id>', methods=['POST'])
def delete_classroom(classroom_id):
    classroom = Classroom.query.get(classroom_id)
    db.session.delete(classroom)
    db.session.commit()
    flash('Classroom deleted successfully.', 'success')
    return redirect(url_for('index'))

@app.route('/add_student', methods=['GET', 'POST'])
def add_student():
    if request.method == 'POST':
        name = request.form['name']
        classroom_id = request.form['classroom_id']
        student = Student(name=name, classroom_id=classroom_id)
        db.session.add(student)
        db.session.commit()
        classroom_id = student.classroom_id
        return redirect('/classroom/{}/students'.format(classroom_id))
    classrooms = Classroom.query.all()
    return render_template('add_student.html', classrooms=classrooms)

@app.route('/remove_student', methods=['POST'])
def remove_student():
    student_id = request.form['student_id']
    classroom_id = request.form['classroom_id']
    student = Student.query.get(student_id)
    for grade in student.grades:
        grade.student_id = None
    db.session.delete(student)
    db.session.commit()
    flash('Student removed successfully.', 'success')
    return redirect('/classroom/{}/students'.format(classroom_id))

@app.route('/add_grade/<int:student_id>', methods=['GET'])
def add_grade(student_id):
    student = Student.query.get(student_id)
    return render_template('add_grade.html', student=student)

@app.route('/insert_grade', methods=['POST'])
def insert_grade():
    if request.method == 'POST':
        grade_value = request.form['grade']
        student_id = request.form['student_id']
        student = Student.query.get(student_id)
        grade = Grade(value=grade_value, student=student)
        db.session.add(grade)
        db.session.commit()
        classroom_id = student.classroom_id
        return redirect('/classroom/{}/students'.format(classroom_id))
    else:
        return "Method not allowed", 405

@app.route('/classroom/<int:classroom_id>/students')
def classroom_students(classroom_id):
    classroom = Classroom.query.get(classroom_id)
    if classroom is None:
        return 'Classroom not found', 404
    students = Student.query.filter_by(classroom_id=classroom_id).all()
    return render_template('classroom_students.html', classroom=classroom, students=students)

@app.route('/export_csv/<int:classroom_id>', methods=['GET'])
def export_csv(classroom_id):
    classroom = Classroom.query.get(classroom_id)
    students = classroom.students

    max_grades = max(len(student.grades) for student in students)

    csv_str = 'Nome'
    for i in range(max_grades):
        csv_str += f',Notas {i+1}'
    csv_str += '\n'
    for student in students:
        grades_str = ','.join([str(grade.value) for grade in student.grades])
        csv_str += f'{student.name},{grades_str}\n'

    response = Response(csv_str, content_type='text/csv')
    response.headers['Content-Disposition'] = 'attachment; filename=Turma_{}_{}h_notas.csv'.format(classroom.name, classroom.time)
    return response

@app.route('/export_json/<int:classroom_id>', methods=['GET'])
def export_json(classroom_id):
    classroom = Classroom.query.get(classroom_id)
    students = classroom.students

    students_data = []
    for student in students:
        grades = [grade.value for grade in student.grades]
        student_data = {
            'nome': student.name,
            'notas': grades
        }
        students_data.append(student_data)

    response_data = {'turma': f'{classroom.name} - {classroom.time}', 'estudantes': students_data}
    response = jsonify(response_data)
    response.headers['Content-Disposition'] = 'attachment; filename=Turma_{}_{}h_notas.json'.format(classroom.name, classroom.time)
    return response

@app.route('/export_yaml/<int:classroom_id>', methods=['GET'])
def export_yaml(classroom_id):
    classroom = Classroom.query.get(classroom_id)
    students = classroom.students

    classroom_data = {
        'turma': f'{classroom.name} - {classroom.time}',
        'estudantes': []
    }

    for student in students:
        student_data = {
            'nome': student.name,
            'notas': [grade.value for grade in student.grades]
        }
        classroom_data['estudantes'].append(student_data)

    yaml_str = yaml.dump(classroom_data, allow_unicode=True)
    response = Response(yaml_str, content_type='text/yaml')
    filename = 'Turma_{}_{}h_notas.yaml'.format(classroom.name, classroom.time)
    response.headers['Content-Disposition'] = 'attachment; filename={}'.format(filename)
    return response

@app.route('/export_xml/<int:classroom_id>', methods=['GET'])
def export_xml(classroom_id):
    classroom = Classroom.query.get(classroom_id)
    students = classroom.students

    xml_str = '<?xml version="1.0" encoding="UTF-8"?>\n'
    xml_str += '<turma>\n'
    for student in students:
        xml_str += '  <estudante>\n'
        xml_str += '    <nome>{}</nome>\n'.format(student.name)
        for grade in student.grades:
            xml_str += '    <nota>{}</nota>\n'.format(grade.value)
        xml_str += '  </estudante>\n'
    xml_str += '</turma>'

    response = Response(xml_str, content_type='application/xml')
    response.headers['Content-Disposition'] = 'attachment; filename=Turma_{}_{}h_notas.xml'.format(classroom.name, classroom.time)
    return response

@app.route('/export_excel/<int:classroom_id>', methods=['GET'])
def export_excel(classroom_id):
    classroom = Classroom.query.get(classroom_id)
    students = classroom.students

    workbook = Workbook()
    sheet = workbook.active

    sheet['A1'] = 'Nome'
    max_grades = 0
    for student in students:
        grades = [str(grade.value) for grade in student.grades]
        max_grades = max(max_grades, len(grades))
    for i in range(max_grades):
        sheet[get_column_letter(i + 2) + '1'] = 'Notas {}'.format(i + 1)

    row = 2
    for student in students:
        sheet['A{}'.format(row)] = student.name
        grades = [str(grade.value) for grade in student.grades]
        for i, grade in enumerate(grades):
            sheet[get_column_letter(i + 2) + str(row)] = grade
        row += 1

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = Response(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response.headers['Content-Disposition'] = 'attachment; filename=Turma_{}_{}h_notas.xlsx'.format(classroom.name, classroom.time)
    return response

if __name__ == '__main__':
    app.run(debug=True)